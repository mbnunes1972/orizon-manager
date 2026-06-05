import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('tabelas_financeiras/xls/tabelas_origem.xlsx', data_only=True)

# Lê aba LEIA-ME
print("=" * 60)
print("ABA: LEIA-ME")
print("=" * 60)
ws = wb['LEIA-ME']
for row in ws.iter_rows():
    vals = [c.value for c in row]
    if any(v is not None for v in vals):
        print('\t'.join(str(v)[:50] if v is not None else '' for v in vals))

# Lê aba aymore com referência às células (com fórmulas se possível)
print()
print("=" * 60)
print("ABA: aymore (com coordenadas)")
print("=" * 60)
wb2 = openpyxl.load_workbook('tabelas_financeiras/xls/tabelas_origem.xlsx', data_only=False)
ws2 = wb2['aymore']
for row in ws2.iter_rows():
    vals = [(c.coordinate, c.value) for c in row]
    non_empty = [(coord, v) for coord, v in vals if v is not None]
    if non_empty:
        for coord, v in non_empty:
            print(f"  {coord}: {str(v)[:80]}")
        print()

# Agora lê o Aymore.xlsx separado
print()
print("=" * 60)
print("ARQUIVO: Aymore.xlsx")
print("=" * 60)
wb3 = openpyxl.load_workbook('tabelas_financeiras/xls/Aymore.xlsx', data_only=True)
print('Abas:', wb3.sheetnames)
for shname in wb3.sheetnames:
    ws3 = wb3[shname]
    print(f"\n--- Aba: {shname} ({ws3.dimensions}) ---")
    for row in ws3.iter_rows():
        vals = [c.value for c in row]
        if any(v is not None for v in vals):
            print('\t'.join(str(v)[:35] if v is not None else '' for v in vals))
