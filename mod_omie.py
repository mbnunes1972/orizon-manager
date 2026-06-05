"""
mod_omie.py — API Omie, exportação, grupos e lógica de clientes.
"""
import os, re, time, json, unicodedata, uuid
import xml.etree.ElementTree as ET
import requests
from datetime import datetime, timedelta
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    openpyxl = None
from storage import (
    _BASE_DIR, PROJETOS_DIR, GRUPOS_CACHE_FILE, CPF_CORINGA,
    storage_ler_json, storage_salvar_json, storage_existe,
    storage_listar, storage_salvar_texto, storage_salvar_binario,
    storage_ler_texto, storage_ler_binario, storage_deletar,
    session_get, session_set,
    _sleep_interval, _omie_key, _omie_secret, _grupos_cache,
    _save_grupos_cache, so_digitos, normalizar, _sha256_str,
    get_omie_key, get_omie_secret, get_sleep_interval, set_sleep_interval
)
from promob_grupos import GRUPOS, ler_xml_str

# == OMIE API LAYER ==
def omie_post(endpoint, call, params, log_cb, no_rotate=False, timeout=10):
    url = "https://app.omie.com.br/api/v1" + endpoint
    for tentativa in range(3):
        payload = {
            "app_key":    get_omie_key(),
            "app_secret": get_omie_secret(),
            "call":       call,
            "param":      [params],
        }
        try:
            resp = requests.post(url, json=payload, timeout=timeout)
            try:
                data = resp.json()
            except Exception:
                data = {}
            if not resp.ok:
                msg = data.get("faultstring") or data.get("message") or resp.text[:300]
                if resp.status_code == 425:
                    raise Exception("OMIE_BLOQUEIO_425: " + msg)
                m = re.search(r"(\d+) segundos", msg)
                if m and tentativa < 2:
                    espera = int(m.group(1))
                    if espera > 300:
                        raise Exception("OMIE_BLOQUEIO_425: bloqueio prolongado (%ds)" % espera)
                    novo = min(espera / 10, 6.0)
                    if novo > get_sleep_interval():
                        set_sleep_interval(novo)
                        log_cb("Rate limit. Aguardando %ds - intervalo ajustado para %.1fs/req" % (espera, novo), "warn")
                    else:
                        log_cb("Rate limit. Aguardando %ds..." % espera, "warn")
                    time.sleep(espera)
                    continue
                raise Exception("HTTP %d: %s" % (resp.status_code, msg))
            if "faultstring" in data:
                raise Exception(data["faultstring"])
            return data
        except Exception as e:
            msg_e = str(e).lower()
            if "connectionreset" in msg_e or "10054" in msg_e or "connection aborted" in msg_e:
                raise Exception("OMIE_BLOQUEIO_425: conexao recusada - conta bloqueada (aguarde 30 min)")
            if "timed out" in msg_e or "timeout" in msg_e:
                raise Exception("OMIE_BLOQUEIO_425: API sem resposta (timeout) - aguarde alguns minutos e tente novamente")
            if tentativa == 2:
                raise
            raise

# == LÓGICA DE PROJETOS ==
def _nome_safe(nome):
    """Remove acentos, substitui espacos por _, remove chars invalidos, max 60."""
    s = ''.join(c for c in unicodedata.normalize('NFD', nome or '') if unicodedata.category(c) != 'Mn')
    s = re.sub(r'[\\/*?:"<>|]', '', s).strip()
    s = re.sub(r'\s+', '_', s)
    return (s or 'projeto')[:60]

def _projeto_path(nome_safe):
    return os.path.join(PROJETOS_DIR, nome_safe)

def _carregar_projeto(nome_safe):
    """Le e retorna projeto.json. Retorna None se nao existir."""
    path = os.path.join(_projeto_path(nome_safe), 'projeto.json')
    if not storage_existe(path):
        return None
    return storage_ler_json(path)

def _salvar_projeto(dados_projeto):
    """Persiste projeto.json atualizando atualizado_em."""
    nome_safe = dados_projeto.get('nome_safe', '')
    pasta = _projeto_path(nome_safe)
    xmls_dir = os.path.join(pasta, 'xmls')
    if not storage_existe(xmls_dir):
        os.makedirs(xmls_dir, exist_ok=True)
    dados_projeto['atualizado_em'] = datetime.now().strftime('%Y-%m-%dT%H:%M:%S')
    storage_salvar_json(os.path.join(pasta, 'projeto.json'), dados_projeto)

def _criar_projeto(nome_projeto, cliente_nome, cliente_cpf='', cliente_email='', cliente_telefone=''):
    """Cria pasta + projeto.json inicial. Retorna o projeto criado."""
    nome_safe = _nome_safe(nome_projeto)
    proj = {
        'nome_safe': nome_safe,
        'nome_projeto': nome_projeto,
        'cliente': {'nome': cliente_nome, 'cpf': cliente_cpf,
                    'email': cliente_email, 'telefone': cliente_telefone},
        'criado_em': datetime.now().strftime('%Y-%m-%dT%H:%M:%S'),
        'atualizado_em': '',
        'codigo_projeto_omie': None,  # preenchido se criação no Omie der certo
        'margens': {
            'desconto_pct': 0.0, 'custo_financeiro_pct': 0.0,
            'custo_viagem': 0.0, 'fora_da_sede': False,
            'brinde': 0.0, 'brinde_ativo': False,
            'comissao_arq_pct': 0.0, 'comissao_arq_ativa': False,
            'fidelidade_pct': 0.0, 'fidelidade_ativa': False
        },
        'forma_pagamento': None,
        'ambientes': [],
        'orcamentos': []
    }
    # Garante unicidade do nome_safe
    base = nome_safe; i = 2
    while storage_existe(_projeto_path(nome_safe)):
        nome_safe = '%s_%d' % (base, i); i += 1
    proj['nome_safe'] = nome_safe
    _salvar_projeto(proj)
    return proj

def _listar_projetos():
    """Lista todos os projetos em PROJETOS/."""
    resultado = []
    for nome_safe in storage_listar(PROJETOS_DIR):
        proj = _carregar_projeto(nome_safe)
        if not proj:
            continue
        # Compatibilidade: app_12 salvava cliente como string
        cli = proj.get('cliente', {})
        if isinstance(cli, str):
            cliente_nome = cli
        else:
            cliente_nome = cli.get('nome', '') if isinstance(cli, dict) else ''
        resultado.append({
            'nome_safe':      nome_safe,
            'nome_projeto':   proj.get('nome_projeto', proj.get('cliente', '')),
            'cliente_nome':   cliente_nome,
            'atualizado_em':  proj.get('atualizado_em', ''),
            'n_ambientes':    len(proj.get('ambientes', [])),
            'n_selecionados': sum(1 for a in proj.get('ambientes', []) if a.get('selecionado')),
        })
    return resultado

def _buscar_projetos(q):
    """Filtra projetos por nome_projeto, cliente.nome ou cliente.cpf."""
    q = q.strip().lower()
    resultado = []
    for p in _listar_projetos():
        if (q in p['nome_projeto'].lower() or
            q in p['cliente_nome'].lower() or
            q in so_digitos(p.get('cliente_cpf', ''))):
            resultado.append(p)
    return resultado

def _buscar_projetos_omie(q):
    """Busca projetos na API do Omie. Retorna [] se falhar ou q vazio."""
    if not q or len(q) < 2:
        return []
    try:
        cfg = config_carregar()
        if not cfg.get('app_key'):
            return []
        _set_credenciais(cfg['app_key'], cfg['app_secret'])
        r = omie_post('/geral/projetos/', 'ListarProjetos',
                      {'pagina': 1, 'registros_por_pagina': 50,
                       'filtrar_por_descricao': q},
                      lambda *_: None, no_rotate=True)
        projetos = []
        for p in r.get('cadastro', []):
            if q.lower() in p.get('cDescricao', '').lower():
                projetos.append({
                    'nome_safe': None,
                    'nome_projeto': p.get('cDescricao', ''),
                    'cliente_nome': '',
                    'atualizado_em': '',
                    'n_ambientes': 0,
                    'n_selecionados': 0,
                    'codigo_omie': p.get('nCodProj'),
                    'origem': 'omie',
                })
        return projetos
    except Exception:
        return []

# == LÓGICA DE AMBIENTES ==
def _extrair_cliente_e_ambiente(xml_str):
    root = ET.fromstring(xml_str)
    cd = {}
    for d in root.findall(".//CUSTOMERSDATA/DATA"):
        vid, val = d.get("ID", ""), d.get("VALUE", "")
        if vid == "nomecliente":
            cd["nome"] = val
        elif vid == "cpfcnpj":
            cd["cpf"] = val
    amb_el = root.find(".//AMBIENT")
    if amb_el is not None:
        cd["_ambiente_desc"] = amb_el.get("DESCRIPTION", "")
    return cd

def carregar_xmls(arquivos_xml):
    ambientes     = []
    cliente_dados = {}
    for arq_nome, arq_conteudo in arquivos_xml:
        try:
            amb   = ler_xml_str(arq_nome, arq_conteudo)
            extra = _extrair_cliente_e_ambiente(arq_conteudo)
            if not cliente_dados.get("nome") and extra.get("nome"):
                cliente_dados = {"nome": extra["nome"], "cpf": extra.get("cpf", "")}
            amb["ambiente"] = extra.get("_ambiente_desc") or amb["projeto"]
            ambientes.append(amb)
        except Exception as e:
            ambientes.append({
                "arquivo": arq_nome, "projeto": arq_nome, "ambiente": arq_nome,
                "data": datetime.today().strftime("%d/%m/%Y"),
                "total": 0.0, "grupos": [], "erro": str(e),
            })
    total_projeto = sum(a["total"] for a in ambientes)
    return {
        "ok":            True,
        "versao":        "3.0",
        "gerado_em":     datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
        "total_projeto": round(total_projeto, 2),
        "cliente":       cliente_dados,
        "ambientes":     ambientes,
        "grupos_ref":    GRUPOS,
    }

def _adicionar_ambientes(nome_safe, arquivos_xml):
    """
    Processa XMLs, copia para xmls/, adiciona/substitui em projeto.json.
    Retorna projeto atualizado.
    """
    projeto = _carregar_projeto(nome_safe)
    if not projeto:
        raise FileNotFoundError('Projeto nao encontrado: %s' % nome_safe)
    pasta_xmls = os.path.join(_projeto_path(nome_safe), 'xmls')
    os.makedirs(pasta_xmls, exist_ok=True)
    for nome_arq, conteudo in arquivos_xml:
        storage_salvar_texto(os.path.join(pasta_xmls, nome_arq), conteudo)
    dados = carregar_xmls(arquivos_xml)
    for amb in dados.get('ambientes', []):
        arq = amb['arquivo']
        projeto['ambientes'] = [a for a in projeto['ambientes'] if a['arquivo'] != arq]
        amb['selecionado'] = True
        amb['arquivo_path'] = os.path.join(pasta_xmls, arq)
        projeto['ambientes'].append(amb)
    _salvar_projeto(projeto)
    return projeto

# == LÓGICA DE GRUPOS OMIE ==
_categoria_cache = [None]
_cc_cache        = [None]

def buscar_cliente_cpf(cpf, log_cb):
    cpf_d  = so_digitos(cpf)
    pagina = 1
    while True:
        try:
            r = omie_post("/geral/clientes/", "ListarClientes",
                          {"pagina": pagina, "registros_por_pagina": 50,
                           "apenas_importado_api": "N"}, log_cb, no_rotate=True)
        except Exception as e:
            msg = str(e).lower()
            if "nao existem registros" in msg:
                return None
            raise
        for c in r.get("clientes_cadastro", []):
            if so_digitos(c.get("cnpj_cpf", "")) == cpf_d:
                return c
        if pagina >= r.get("total_de_paginas", 1):
            break
        pagina += 1
    return None

def pesquisar_clientes(query, por_email=False):
    query  = query.strip()
    cpf_q  = so_digitos(query)
    norm_q = normalizar(query)
    silent = lambda *_: None
    resultados = []
    pagina = 1
    max_pag = 10 if cpf_q else 5
    while pagina <= max_pag:
        try:
            r = omie_post("/geral/clientes/", "ListarClientes",
                          {"pagina": pagina, "registros_por_pagina": 50,
                           "apenas_importado_api": "N"}, silent)
        except Exception:
            break
        for c in r.get("clientes_cadastro", []):
            match = False
            if cpf_q and so_digitos(c.get("cnpj_cpf", "")) == cpf_q:
                match = True
            elif por_email and query.lower() == (c.get("email") or "").lower():
                match = True
            elif not cpf_q and not por_email and norm_q and norm_q in normalizar(c.get("razao_social", "")):
                match = True
            if match:
                resultados.append(c)
            if len(resultados) >= 15:
                break
        if len(resultados) >= 15 or pagina >= r.get("total_de_paginas", 1):
            break
        pagina += 1
    return resultados

def criar_cliente(nome, cpf, log_cb):
    slug       = re.sub(r"[^A-Z0-9]", "", normalizar(nome))[:12]
    integracao = ("CORINGA_" + slug) if so_digitos(cpf) == so_digitos(CPF_CORINGA) else slug
    r = omie_post("/geral/clientes/", "IncluirCliente", {
        "codigo_cliente_integracao": integracao,
        "razao_social":    nome,
        "nome_fantasia":   nome,
        "pessoa_fisica":   "S",
        "inativo":         "N",
        "cnpj_cpf":        cpf,
        "email":           "",
        "endereco":        "",
        "endereco_numero": "",
        "bairro":          "",
        "cidade":          "",
        "estado":          "",
        "cep":             "",
        "telefone1_ddd":    "00",
        "telefone1_numero": "000000000",
    }, log_cb, no_rotate=True, timeout=30)
    return r["codigo_cliente_omie"]

def garantir_conta_corrente(log_cb):
    if _cc_cache[0] is not None:
        return _cc_cache[0]
    try:
        r = omie_post("/geral/contacorrente/", "ListarContasCorrentes",
                      {"pagina": 1, "registros_por_pagina": 50}, log_cb)
        for c in r.get("ListarContasCorrentes", []):
            if c.get("inativo", "N") == "N":
                cod  = c.get("nCodCC")
                nome = c.get("descricao", "")
                if cod:
                    _cc_cache[0] = cod
                    log_cb("  Conta corrente: %s (#%s)" % (nome, cod), "ok")
                    return cod
    except Exception as e:
        log_cb("  Conta corrente nao obtida: %s" % e, "warn")
    return None

def buscar_categoria(log_cb):
    if _categoria_cache[0]:
        return _categoria_cache[0]
    try:
        r    = omie_post("/geral/categorias/", "ListarCategorias",
                         {"pagina": 1, "registros_por_pagina": 100}, log_cb)
        cats = r.get("categoria_cadastro", [])
        for cat in cats:
            cod = cat.get("codigo", "")
            if re.match(r"^1[.]\d+[.]\d+$", cod) and "<" not in cat.get("descricao", ""):
                _categoria_cache[0] = cod
                log_cb("  Categoria: %s - %s" % (cod, cat.get("descricao", "")), "ok")
                return cod
        for cat in cats:
            cod = cat.get("codigo", "")
            if re.match(r"^\d+[.]\d+[.]\d+$", cod) and "<" not in cat.get("descricao", ""):
                _categoria_cache[0] = cod
                log_cb("  Categoria (fallback): %s - %s" % (cod, cat.get("descricao", "")), "ok")
                return cod
    except Exception as e:
        log_cb("  Nao foi possivel obter categoria: %s" % e, "warn")
    return None

def criar_pedido(codigo_cliente, itens, obs, data_prev, log_cb):
    cod_integracao = "PROMOB-" + uuid.uuid4().hex[:16].upper()
    det = [
        {
            "ide": {"codigo_item_integracao": "ITEM%02d" % (idx + 1)},
            "produto": {
                "codigo_produto":      i["codigo_produto"],
                "quantidade":          i["quantidade"],
                "valor_unitario":      i["valor_unitario"],
                "tipo_desconto":       "V",
                "percentual_desconto": 0,
            },
        }
        for idx, i in enumerate(itens)
    ]
    categoria = buscar_categoria(log_cb)
    cc        = garantir_conta_corrente(log_cb)
    inf_adic  = {"utilizar_emails": "N"}
    if categoria:
        inf_adic["codigo_categoria"]      = categoria
    if cc:
        inf_adic["codigo_conta_corrente"] = cc
    r = omie_post("/produtos/pedido/", "IncluirPedido", {
        "cabecalho": {
            "codigo_cliente":           codigo_cliente,
            "data_previsao":            data_prev,
            "etapa":                    "10",
            "codigo_pedido_integracao": cod_integracao,
        },
        "det":                    det,
        "informacoes_adicionais": inf_adic,
    }, log_cb, no_rotate=True, timeout=60)
    return r.get("numero_pedido", r.get("codigo_pedido", "?"))

def _consultar_produto(ref, log_cb=None):
    r = omie_post("/geral/produtos/", "ConsultarProduto",
                  {"codigo_produto_integracao": ref},
                  log_cb or (lambda *_: None), no_rotate=True)
    return (r.get("produto_servico_cadastro") or {}).get("codigo_produto") or r.get("codigo_produto")

def garantir_grupos_omie(log_cb):
    grupos_omie   = {}
    refs_esperadas = ["GRUPO-" + g["id"] for g in GRUPOS]
    if all(ref in _grupos_cache for ref in refs_esperadas):
        for g in GRUPOS:
            grupos_omie[g["id"]] = _grupos_cache["GRUPO-" + g["id"]]
        log_cb("16/16 grupos carregados do cache local (sem chamadas a API)", "ok")
        return grupos_omie
    log_cb("Verificando produtos fixos GRUPO-01 a GRUPO-16...", "section")
    for g in GRUPOS:
        ref = "GRUPO-" + g["id"]
        if ref in _grupos_cache:
            grupos_omie[g["id"]] = _grupos_cache[ref]
            log_cb("  OK %s (cache)" % ref, "ok")
            continue
        log_cb("  -> Consultando %s..." % ref, "info")
        time.sleep(_sleep_interval[0])
        try:
            cod = _consultar_produto(ref, log_cb)
            if cod:
                _grupos_cache[ref]    = cod
                grupos_omie[g["id"]] = cod
                _save_grupos_cache(_grupos_cache)
                log_cb("  OK %s - %s" % (ref, g["nome"][:50]), "ok")
        except Exception as e:
            msg = str(e).lower()
            if "nao cadastrado" in msg or "nao encontrado" in msg:
                for ncm in [g["ncm"], "94036000"]:
                    time.sleep(_sleep_interval[0])
                    try:
                        r2 = omie_post("/geral/produtos/", "IncluirProduto", {
                            "codigo_produto_integracao": ref,
                            "codigo":         ref,
                            "descricao":      g["nome"][:120],
                            "unidade":        "VB",
                            "ncm":            ncm,
                            "tipoItem":       "00",
                            "valor_unitario": 1.00,
                        }, log_cb, no_rotate=True)
                        cod = r2["codigo_produto"]
                        _grupos_cache[ref]    = cod
                        grupos_omie[g["id"]] = cod
                        _save_grupos_cache(_grupos_cache)
                        sufixo = " (NCM 94036000)" if ncm == "94036000" and ncm != g["ncm"] else ""
                        log_cb("  CRIADO %s%s - %s" % (ref, sufixo, g["nome"][:50]), "create")
                        break
                    except Exception as e2:
                        msg2 = str(e2).lower()
                        if "ja cadastrado" in msg2:
                            try:
                                cod = _consultar_produto(ref, log_cb)
                                if cod:
                                    _grupos_cache[ref]    = cod
                                    grupos_omie[g["id"]] = cod
                                    _save_grupos_cache(_grupos_cache)
                                    log_cb("  %s ja existia, codigo recuperado" % ref, "ok")
                            except Exception:
                                log_cb("  AVISO %s: nao foi possivel recuperar codigo" % ref, "warn")
                            break
                        elif "ncm" in msg2 and ncm != "94036000":
                            log_cb("  %s: NCM %s rejeitada - retentando com 94036000" % (ref, ncm), "warn")
                            continue
                        else:
                            log_cb("  AVISO %s: %s" % (ref, e2), "warn")
                            break
            elif "omie_bloqueio_425" in msg:
                log_cb("  BLOQUEADO %s - API bloqueada (aguarde ~30 min)" % ref, "err")
                faltando = ["GRUPO-" + x["id"] for x in GRUPOS if x["id"] not in grupos_omie]
                if faltando:
                    log_cb("  Pulando %d grupos: %s" % (len(faltando), ", ".join(faltando)), "warn")
                break
            else:
                log_cb("  AVISO %s: %s" % (ref, e), "warn")
    ok = len(grupos_omie)
    log_cb("  %d/16 grupos verificados no Omie" % ok, "ok" if ok == 16 else "warn")
    return grupos_omie

def exportar_ambientes(app_key, app_secret, dados, ambientes_sel, log_cb, confirm_cb, intervalo=None):
    if intervalo is not None:
        _sleep_interval[0] = float(intervalo)
    _categoria_cache[0] = None
    _cc_cache[0]        = None
    todos = dados.get("ambientes", [])
    if not ambientes_sel or ambientes_sel == "todos":
        selecionados = todos
    else:
        selecionados = [a for a in todos if a["arquivo"] in ambientes_sel]
    if not selecionados:
        log_cb("Nenhum ambiente selecionado para exportar.", "warn")
        log_cb("__DONE__", "done")
        return []
    grupos_omie = garantir_grupos_omie(log_cb)
    if not grupos_omie:
        log_cb("Nenhum grupo disponivel no Omie - abortando.", "err")
        log_cb("__DONE__", "done")
        return []
    total_geral = sum(a["total"] for a in selecionados)
    log_cb("Total selecionado: R$ %s" % "{:,.2f}".format(total_geral), "info")
    cli_sel = session_get("cliente_selecionado")
    if cli_sel:
        codigo_cliente = cli_sel["codigo"]
        log_cb("Cliente: %s (#%s)" % (cli_sel["nome"], codigo_cliente), "info")
    else:
        nome_cliente = dados.get("cliente", {}).get("nome", "CLIENTE")
        cpf_xml      = so_digitos(dados.get("cliente", {}).get("cpf", ""))
        log_cb("Cliente (XML): %s" % nome_cliente, "info")
        eh_coringa = False
        cpf_final  = cpf_xml or so_digitos(CPF_CORINGA)
        if not cpf_xml:
            cpf_input = confirm_cb("sem_cpf", {"nome": nome_cliente})
            cpf_final = so_digitos(cpf_input) if cpf_input else so_digitos(CPF_CORINGA)
            if not so_digitos(cpf_input or ""):
                eh_coringa = True
                log_cb("Usando CPF coringa - atualizar no contrato", "warn")
        cliente_omie = buscar_cliente_cpf(cpf_final, log_cb)
        if cliente_omie:
            codigo_cliente = cliente_omie["codigo_cliente_omie"]
            nome_omie      = cliente_omie["razao_social"]
            if normalizar(nome_omie) != normalizar(nome_cliente):
                acao = confirm_cb("nome_diferente", {"nome_xml": nome_cliente, "nome_omie": nome_omie})
                if acao == "atualizar":
                    omie_post("/geral/clientes/", "AlterarCliente",
                              {"codigo_cliente_omie": codigo_cliente,
                               "razao_social": nome_cliente,
                               "nome_fantasia": nome_cliente}, log_cb, no_rotate=True)
                    log_cb("Nome atualizado para '%s'" % nome_cliente, "ok")
                else:
                    log_cb("Usando nome do Omie: '%s'" % nome_omie, "ok")
            else:
                log_cb("Cliente encontrado: %s (#%s)" % (nome_omie, codigo_cliente), "ok")
            if not cliente_omie.get("estado", "").strip():
                uf = confirm_cb("sem_uf", {"nome": nome_omie})
                uf = (uf or "").strip().upper()[:2]
                if uf:
                    omie_post("/geral/clientes/", "AlterarCliente",
                              {"codigo_cliente_omie": codigo_cliente, "estado": uf},
                              log_cb, no_rotate=True, timeout=30)
                    log_cb("UF '%s' adicionada ao cadastro do cliente" % uf, "ok")
        else:
            uf_input = confirm_cb("sem_uf", {"nome": nome_cliente})
            uf       = (uf_input or "").strip().upper()[:2]
            log_cb("  -> Criando cliente '%s'..." % nome_cliente, "info")
            codigo_cliente = criar_cliente(
                nome_cliente,
                cpf_final if not eh_coringa else CPF_CORINGA,
                log_cb,
            )
            if uf:
                omie_post("/geral/clientes/", "AlterarCliente",
                          {"codigo_cliente_omie": codigo_cliente, "estado": uf},
                          log_cb, no_rotate=True, timeout=30)
            log_cb("Cliente criado (#%s)" % codigo_cliente, "ok")

    def _itens_inteiros(itens, cod_grupo16):
        ajustados = []
        residuo   = 0.0
        for item in itens:
            qty_int  = int(item["quantidade"])
            residuo += item["quantidade"] - qty_int
            if qty_int > 0:
                ajustados.append(dict(item, quantidade=float(qty_int)))
        residuo_round = round(residuo)
        if residuo_round > 0 and cod_grupo16:
            g16 = next((i for i in ajustados if i["codigo_produto"] == cod_grupo16), None)
            if g16:
                g16["quantidade"] += residuo_round
            else:
                ajustados.append({"codigo_produto": cod_grupo16, "quantidade": float(residuo_round), "valor_unitario": 1.00})
        return ajustados, round(residuo, 2)

    pedidos = []
    for amb in selecionados:
        nome_amb   = amb.get("ambiente") or amb.get("projeto") or amb["arquivo"]
        grupos_amb = amb.get("grupos", [])
        if not grupos_amb:
            log_cb("\nAMBIENTE %s - sem grupos, pulando" % nome_amb, "warn")
            continue
        log_cb("\nAmbiente: %s  (%d grupos | R$ %s)" % (
            nome_amb, len(grupos_amb), "{:,.2f}".format(amb["total"])), "section")
        itens_omie = []
        for g in grupos_amb:
            gid = g["grupo_id"]
            if gid not in grupos_omie:
                log_cb("  AVISO GRUPO-%s indisponivel - linha omitida" % gid, "warn")
                continue
            itens_omie.append({"codigo_produto": grupos_omie[gid], "quantidade": g["subtotal"], "valor_unitario": 1.00})
        if not itens_omie:
            log_cb("  Nenhum grupo com produto disponivel - pedido nao criado", "warn")
            continue
        total_real = sum(i["quantidade"] for i in itens_omie)
        time.sleep(2.0)
        data_prev = amb.get("data") or datetime.today().strftime("%d/%m/%Y")
        try:
            num = criar_pedido(codigo_cliente, itens_omie, nome_amb, data_prev, log_cb)
        except Exception as e_ped:
            msg_e = str(e_ped).lower()
            if any(kw in msg_e for kw in ("quantidade", "decimal", "fracion", "numero invalido")):
                log_cb("  Quantidade decimal rejeitada - aplicando ajuste", "warn")
                cod16  = grupos_omie.get("16")
                itens_int, residuo = _itens_inteiros(itens_omie, cod16)
                if residuo > 0:
                    log_cb("  Ajuste: R$ %.2f incorporados em GRUPO-16 (COMPLEMENTOS)" % residuo, "warn")
                num = criar_pedido(codigo_cliente, itens_int, nome_amb, data_prev, log_cb)
            else:
                raise
        pedidos.append({"ambiente": nome_amb, "numero": num, "total": round(total_real, 2)})
        log_cb("Pedido #%s criado - %d linha(s) - R$ %s" % (
            num, len(itens_omie), "{:,.2f}".format(total_real)), "ok")
    log_cb("\nResumo: %d pedido(s) criado(s)" % len(pedidos), "section")
    log_cb("__DONE__", "done")
    return pedidos

# -- Exportacao Excel -----------------------------------------------------------
def gerar_excel(dados, nome_cliente=None):
    """Retorna (bytes, display_name, caminho_completo)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise Exception("Modulo openpyxl nao instalado. Execute: pip install openpyxl")

    COR_TITULO = "1A1D27"
    COR_HEADER = "FF6B35"
    COR_TOTAL  = "D6F5EF"
    COR_GRUPO  = "252836"
    thin  = Side(style="thin", color="252836")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    def estilo_celula(ws, cel, valor, negrito=False, cor_fonte="E8EAF0",
                      cor_fundo=None, alinhamento="left", borda_=True):
        c = ws[cel]
        c.value = valor
        c.font  = Font(name="Calibri", size=10, bold=negrito, color=cor_fonte)
        if cor_fundo:
            c.fill = PatternFill(fill_type="solid", fgColor=cor_fundo)
        c.alignment = Alignment(horizontal=alinhamento, vertical="center")
        if borda_:
            c.border = borda

    def adicionar_aba(wb, amb):
        nome_aba = re.sub(r"[\\/*?:\[\]]", " ", amb.get("ambiente") or amb.get("projeto") or "Amb")[:31]
        if nome_aba in wb.sheetnames:
            del wb[nome_aba]
        ws = wb.create_sheet(title=nome_aba)
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 36
        ws.column_dimensions["C"].width = 16
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 22
        estilo_celula(ws, "A1", "Promob -> Omie", negrito=True, cor_fonte="FFFFFF",
                      cor_fundo=COR_TITULO, alinhamento="left", borda_=False)
        ws.merge_cells("A1:C1")
        estilo_celula(ws, "A2", amb.get("ambiente") or amb.get("projeto"), negrito=True,
                      cor_fonte="FF6B35", cor_fundo=COR_TITULO, borda_=False)
        ws.merge_cells("A2:C2")
        ws.row_dimensions[3].height = 18
        for col, txt in [("A3", "GRUPO"), ("B3", "DESCRICAO"), ("C3", "SUBTOTAL (R$)")]:
            estilo_celula(ws, col, txt, negrito=True, cor_fonte="FFFFFF",
                          cor_fundo=COR_HEADER, alinhamento="center")
        linha = 4
        total_amb = 0.0
        for g in amb.get("grupos", []):
            ws.row_dimensions[linha].height = 16
            estilo_celula(ws, "A%d" % linha, "GRUPO-%s" % g["grupo_id"], cor_fundo=COR_GRUPO, alinhamento="center")
            estilo_celula(ws, "B%d" % linha, g["grupo_nome"], cor_fundo=COR_GRUPO)
            estilo_celula(ws, "C%d" % linha, g["subtotal"], cor_fundo=COR_GRUPO, alinhamento="right")
            ws["C%d" % linha].number_format = "#,##0.00"
            total_amb += g["subtotal"]
            linha += 1
        ws.row_dimensions[linha].height = 20
        estilo_celula(ws, "A%d" % linha, "TOTAL", negrito=True, cor_fundo=COR_TOTAL, alinhamento="center", cor_fonte="1A1D27")
        estilo_celula(ws, "B%d" % linha, "", cor_fundo=COR_TOTAL, cor_fonte="1A1D27")
        estilo_celula(ws, "C%d" % linha, round(total_amb, 2), negrito=True, cor_fundo=COR_TOTAL, alinhamento="right", cor_fonte="1A1D27")
        ws["C%d" % linha].number_format = "#,##0.00"

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    for amb in dados.get("ambientes", []):
        adicionar_aba(wb, amb)

    if not wb.sheetnames:
        ws = wb.create_sheet("Sem dados")
        ws["A1"] = "Nenhum ambiente com grupos"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    conteudo = buf.read()

    nome_saida = "orcamento_%s.xlsx" % datetime.now().strftime("%Y%m%d_%H%M")
    caminho    = os.path.join(_BASE_DIR, nome_saida)

    storage_salvar_binario(caminho, conteudo)

    return conteudo, nome_saida, caminho

# -- Arquivar XMLs --------------------------------------------------------------
def _arquivar_xmls(nome_safe, arquivos_xml):
    """Copia XMLs para pasta de historico com timestamp."""
    pasta = os.path.join(PROJETOS_DIR, nome_safe)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    pasta_hist = os.path.join(pasta, "xmls_aprovados", ts)
    os.makedirs(pasta_hist, exist_ok=True)
    for nome_arq, conteudo in arquivos_xml:
        storage_salvar_texto(os.path.join(pasta_hist, nome_arq), conteudo)
    return pasta_hist
